# ======================================================================
# GenerateAnalysis.py  (FINAL CLEAN VERSION)
# Correct SHORT Logic + Correct Date Order + MATRIX_MAPPING Restored
# ======================================================================

import os
import glob
import warnings
import numpy as np
import pandas as pd
import configparser

from config_loader import load_config  # type: ignore

# ======================================================================
# CONFIG
# ======================================================================

cfg = load_config()

TARGET_DIRECTORY = cfg.get("TARGET_DIRECTORY")
SD_MULTIPLIER = float(cfg.get("SD_MULTIPLIER", 1.0))
TRADING_QTY = cfg.get("TRADING_QTY", 0)
SYMBOL = cfg.get("SYMBOL", "SYMBOL")
CONFIG_PATH = cfg.get("CONFIG_PATH", None)

if CONFIG_PATH is None:
    CONFIG_PATH = os.path.join(os.getcwd(), "configProcess.ini")

print("Using configuration (GenerateAnalysis.py):")
print(" TARGET_DIRECTORY =", TARGET_DIRECTORY)
print(" SD_MULTIPLIER   =", SD_MULTIPLIER)
print(" TRADING_QTY     =", TRADING_QTY)
print(" SYMBOL          =", SYMBOL)
print(" CONFIG_PATH     =", CONFIG_PATH)

# ======================================================================
# CONSTANTS
# ======================================================================

EQUITY_CLOSE_PRICE_COL = "close"
VWAP_COL = "vwap"
DELIVERY_QTY_RAW_COL_CLEANED = "Deliverable_Qty"
DELIVERY_QTY_FINAL_COL = "Daily_Deliverable_Qty"
DELIVERY_VALUE_COL = "Delivery"
NEW_5DAD_COL = "5DAD"
REL_DELIVERY_COL = "~Del"
PRICE_CHANGE_COL = "~Price"
OI_CHANGE_COL = "~OI"

DATE_COL_CANDIDATES = [
    "DATE", "DATE_", "TRADING_DATE", "TRADE_DATE",
    "TIMESTAMP", "DATES", "Date"
]

# ======================================================================
# MATRIX MAPPING (F&O Conclusion)
# ======================================================================

MATRIX_MAPPING = {
    (1, 1, 1): ("StrongLong", "BUY"),
    (1, 1, 0): ("LastLegOfLong", "WEAK_BUY"),
    (1, 1, -1): ("ShortCovering", "BUY"),
    (1, 0, 1): ("NewLongs", "BUY"),
    (1, 0, 0): ("NoInterest", "NO_TRADE"),
    (1, 0, -1): ("WeakShortCovering", "WEAK_BUY"),
    (1, -1, 1): ("WeakerLongs", "NO_TRADE"),
    (1, -1, 0): ("NoLongPosition", "NO_TRADE"),
    (1, -1, -1): ("WeakShortcovering", "NO_TRADE"),

    (-1, 1, 1): ("StrongShort", "SELL"),
    (-1, 1, 0): ("LastLegOfshort", "WEAK_SELL"),
    (-1, 1, -1): ("LongCovering", "SELL"),
    (-1, 0, 1): ("NewShort", "SELL"),
    (-1, 0, 0): ("NoInterest", "NO_TRADE"),
    (-1, 0, -1): ("WeakLongcovering", "WEAK_SELL"),
    (-1, -1, 1): ("WeakerShort", "WEAK_SELL"),
    (-1, -1, 0): ("NoShortPosition", "NO_TRADE"),
    (-1, -1, -1): ("WeakLongCovering", "WEAK_SELL"),
}

# ======================================================================
# CLEAN NUMERIC
# ======================================================================

def clean_numeric(series):
    return (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace(r"[^\d\.\-]", "", regex=True)
        .str.strip()
        .replace("-", "0")
        .replace("", "0")
        .astype(float)
        .fillna(0.0)
    )

# ======================================================================
# CLEAN DATAFRAME
# ======================================================================

def _clean_dataframe(df, date_candidates, new_date="DATE", dayfirst=True):
    df.columns = (
        df.columns.str.strip()
        .str.replace(" ", "_")
        .str.replace('"', "")
        .str.replace(".", "")
    )
    df = df.loc[:, ~df.columns.duplicated()]

    found = next(
        (c for c in df.columns if c.upper() in [x.upper() for x in date_candidates]),
        None
    )
    if found and found != new_date:
        df.rename(columns={found: new_date}, inplace=True)

    if new_date in df.columns:
        df[new_date] = pd.to_datetime(df[new_date], errors="coerce", dayfirst=dayfirst)
        df.dropna(subset=[new_date], inplace=True)

    return df

# ======================================================================
# DIRECTIONAL SIGNAL USING SD
# ======================================================================

def get_directional_signal_with_sd(series, multiplier):
    s = series.copy()
    valid = s.dropna()

    if len(valid) < 10:
        return pd.Series(0, index=s.index).astype(int), 0.0

    threshold = valid.std() * multiplier

    def classify(v):
        if v > threshold:
            return 1
        if v < -threshold:
            return -1
        return 0

    return s.apply(classify).fillna(0).astype(int), threshold

# ======================================================================
# BUILD BASE DATAFRAME
# ======================================================================

def build_base_dataframe(target_directory, sd_multiplier):

    # ------------------- EQUITY -------------------
    eq_files = glob.glob(os.path.join(target_directory, "Quote-Equity-*.csv"))
    if not eq_files:
        raise FileNotFoundError("No equity files found")

    eq_list = []
    for fn in eq_files:
        df = pd.read_csv(fn, encoding="utf-8-sig")
        df = _clean_dataframe(df, DATE_COL_CANDIDATES)
        for c in df.columns:
            if df[c].dtype == "object":
                df[c] = df[c].str.replace(",", "")
        df = df.apply(pd.to_numeric, errors="ignore")
        eq_list.append(df)

    equity_df = pd.concat(eq_list, ignore_index=True)

    # ------------------- DELIVERY -------------------
    del_files = glob.glob(os.path.join(target_directory, "*-EQ-N.csv"))
    if del_files:
        del_list = []
        for fn in del_files:
            df = pd.read_csv(fn, encoding="utf-8-sig")
            df = _clean_dataframe(df, DATE_COL_CANDIDATES)
            for c in df.columns:
                if df[c].dtype == "object":
                    df[c] = df[c].str.replace(",", "")
            df = df.apply(pd.to_numeric, errors="ignore")
            del_list.append(df)

        delivery_df = pd.concat(del_list, ignore_index=True)

        if DELIVERY_QTY_RAW_COL_CLEANED in delivery_df.columns:
            delivery_df[DELIVERY_QTY_FINAL_COL] = delivery_df[DELIVERY_QTY_RAW_COL_CLEANED]
        else:
            delivery_df[DELIVERY_QTY_FINAL_COL] = 0

        delivery_df = delivery_df[["DATE", DELIVERY_QTY_FINAL_COL]]
    else:
        delivery_df = pd.DataFrame(columns=["DATE", DELIVERY_QTY_FINAL_COL])

    # ------------------- F&O -------------------
    fao_files = glob.glob(os.path.join(target_directory, "*FAO*.csv"))
    if fao_files:
        fao_list = []
        for fn in fao_files:
            df = pd.read_csv(fn, encoding="utf-8-sig")
            df = _clean_dataframe(df, DATE_COL_CANDIDATES)
            fao_list.append(df)
        fao = pd.concat(fao_list, ignore_index=True)

        for c in ["Volume", "OPEN_INTEREST"]:
            if c in fao.columns:
                fao[c] = clean_numeric(fao[c])

        aggregated = (
            fao.groupby("DATE")
            .agg({"Volume": "sum", "OPEN_INTEREST": "sum"})
            .reset_index()
        )
        aggregated.rename(columns={
            "Volume": "Daily_F&O_Volume_Sum",
            "OPEN_INTEREST": "Daily_Open_Interest_Sum"
        }, inplace=True)
    else:
        aggregated = pd.DataFrame(
            columns=["DATE", "Daily_F&O_Volume_Sum", "Daily_Open_Interest_Sum"]
        )

    # ------------------- MERGE -------------------
    for df in [equity_df, delivery_df, aggregated]:
        if "DATE" in df.columns:
            df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce")

    df = equity_df.merge(aggregated, on="DATE", how="left")
    df = df.merge(delivery_df, on="DATE", how="left")

    df["Daily_Open_Interest_Sum"] = clean_numeric(df.get("Daily_Open_Interest_Sum", 0))
    df[DELIVERY_QTY_FINAL_COL] = df[DELIVERY_QTY_FINAL_COL].fillna(0)

    # ------------------- CLEAN VWAP / CLOSE -------------------
    if VWAP_COL in df.columns:
        df[VWAP_COL] = clean_numeric(df[VWAP_COL]).fillna(0)
    else:
        df[VWAP_COL] = 0

    if EQUITY_CLOSE_PRICE_COL not in df.columns:
        for alt in ["Close", "close", "ltp", "LastPrice"]:
            if alt in df.columns:
                df.rename(columns={alt: EQUITY_CLOSE_PRICE_COL}, inplace=True)
                break
    df[EQUITY_CLOSE_PRICE_COL] = clean_numeric(df[EQUITY_CLOSE_PRICE_COL])

    # ------------------- PRICE CHANGE -------------------
    prev_close = df[EQUITY_CLOSE_PRICE_COL].shift(1)
    df[PRICE_CHANGE_COL] = (
        (df[EQUITY_CLOSE_PRICE_COL] - prev_close) /
        prev_close.replace(0, np.nan) * 100
    )

    # ------------------- DELIVERY VALUE -------------------
    del_qty = clean_numeric(df[DELIVERY_QTY_FINAL_COL])
    vwap_clean = clean_numeric(df[VWAP_COL])
    df["Del_Inter"] = del_qty * vwap_clean
    df[DELIVERY_VALUE_COL] = df["Del_Inter"] / 10000000

    # ------------------- OI CHANGE -------------------
    prev_oi = df["Daily_Open_Interest_Sum"].shift(1)
    df[OI_CHANGE_COL] = (
        (df["Daily_Open_Interest_Sum"] - prev_oi) /
        prev_oi.replace(0, 1e-6) * 100
    )
    df["Absolute_OI_Change"] = df["Daily_Open_Interest_Sum"] - prev_oi

    # ------------------- 5D AVERAGE DELIVERY -------------------
    df[NEW_5DAD_COL] = (
        df[DELIVERY_VALUE_COL].shift(1).rolling(window=5, min_periods=5).mean()
    )

    df[REL_DELIVERY_COL] = np.where(
        df[NEW_5DAD_COL] != 0,
        (df[DELIVERY_VALUE_COL] / df[NEW_5DAD_COL]) * 100,
        np.nan,
    )

    # ------------------- DIRECTIONAL SIGNALS -------------------
    df["Price_Dir"], _ = get_directional_signal_with_sd(df[PRICE_CHANGE_COL], sd_multiplier)
    df["Delivery_Dir"], _ = get_directional_signal_with_sd(df[REL_DELIVERY_COL], sd_multiplier)
    df["OI_Dir"], _     = get_directional_signal_with_sd(df[OI_CHANGE_COL], sd_multiplier)

    df["Scenario_Tuple"] = list(zip(
        df["Price_Dir"],
        df["Delivery_Dir"],
        df["OI_Dir"]
    ))

    df["F&O_Conclusion"] = df["Scenario_Tuple"].apply(
        lambda x: MATRIX_MAPPING.get(x, ("Unknown", "NO_TRADE"))[0]
    )

    return df.reset_index(drop=True)

# ======================================================================
# LOAD THRESHOLDS
# ======================================================================

def load_thresholds_from_config():
    parser = configparser.ConfigParser()
    if not os.path.exists(CONFIG_PATH):
        return {}

    parser.read(CONFIG_PATH)
    section = f"THRESHOLDS_{SYMBOL}"

    if not parser.has_section(section):
        return {}

    out = {}
    for k, v in parser.items(section):
        try:
            out[k] = float(v)
        except:
            pass

    return out

# ======================================================================
# APPLY THRESHOLDS (with Correct Short Logic + Correct Date Ordering)
# ======================================================================

def apply_thresholds_and_generate_files(target_directory, sd_multiplier):

    df = build_base_dataframe(target_directory, sd_multiplier)

    # ------------------- IMPORTANT: SORT ASCENDING FOR CUMSUM -------------------
    df = df.sort_values("DATE").reset_index(drop=True)

    thr = load_thresholds_from_config()

    price_long  = thr.get("price_long")
    del_long    = thr.get("del_long")
    oi_long     = thr.get("oi_long")

    price_short = thr.get("price_short")
    del_short   = thr.get("del_short")
    oi_short    = thr.get("oi_short")

    # ------------------- LONG CONDITIONS -------------------
    df["Above_Price_Thr"] = df[PRICE_CHANGE_COL] > price_long
    df["Above_Del_Thr"]   = df[REL_DELIVERY_COL] > del_long
    df["Above_OI_Thr"]    = df[OI_CHANGE_COL] > oi_long

    df["LONG_TRIGGER"] = (
        df["Above_Price_Thr"] &
        df["Above_Del_Thr"] &
        df["Above_OI_Thr"]
    )

    # ------------------- SHORT CONDITIONS (Correct Logic) -------------------
    df["Below_Price_Thr"] = df[PRICE_CHANGE_COL] < price_short
    df["Below_OI_Thr"]    = df[OI_CHANGE_COL] < oi_short

    df["SHORT_TRIGGER"] = (
        df["Below_Price_Thr"] &
        df["Below_OI_Thr"] &
        (df[REL_DELIVERY_COL] > del_short)   # <-- Correct short logic
    )

    # ------------------- LONGS & SHORTS (Absolute OI Change) -------------------
    df["Longs"]  = 0.0
    df["Shorts"] = 0.0

    df.loc[df["LONG_TRIGGER"],  "Longs"]  = df.loc[df["LONG_TRIGGER"], "Absolute_OI_Change"].abs()
    df.loc[df["SHORT_TRIGGER"], "Shorts"] = df.loc[df["SHORT_TRIGGER"], "Absolute_OI_Change"].abs()

    # ------------------- CUMSUM (Correct Now Due to Ascending Sort) -------------------
    df["Longs Till Now"]  = df["Longs"].cumsum()
    df["Shorts Till Now"] = df["Shorts"].cumsum()

    df["Date Display (dd-MM-yyyy)"] = df["DATE"].dt.strftime("%d-%m-%Y")
    df[" "] = np.nan

    # ------------------- OUTPUT ORDER -------------------
    tail_cols = [
        "Del_Inter", DELIVERY_VALUE_COL, NEW_5DAD_COL, " ",
        PRICE_CHANGE_COL, REL_DELIVERY_COL, OI_CHANGE_COL,
        "Absolute_OI_Change",
        "Longs", "Shorts", "Longs Till Now", "Shorts Till Now",
        "Price_Dir", "Delivery_Dir", "OI_Dir",
        "Scenario_Tuple", "F&O_Conclusion",
        "Above_Price_Thr", "Above_Del_Thr", "Above_OI_Thr",
        "Below_Price_Thr", "Below_OI_Thr",
        "LONG_TRIGGER", "SHORT_TRIGGER",
        "Date Display (dd-MM-yyyy)"
    ]

    all_cols = df.columns.tolist()
    first_cols = [c for c in all_cols if c not in tail_cols]
    df = df.loc[:, first_cols + tail_cols]

    # ------------------- FINAL OUTPUT ORDER (DESCENDING DISPLAY) -------------------
    df = df.sort_values("DATE", ascending=True).reset_index(drop=True)

    # ------------------- SAVE CSV -------------------
    csv_path = os.path.join(target_directory, f"{SYMBOL}_Analysis.csv")
    df.to_csv(csv_path, index=False)
    print("Saved CSV:", csv_path)

    # ------------------- SAVE EXCEL -------------------
    xlsx_path = os.path.join(target_directory, f"{SYMBOL}_Analysis_Excel.xlsx")
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Analysis")

            from openpyxl.styles import PatternFill, Font
            ws = writer.sheets["Analysis"]

            header_fill = PatternFill(start_color="0000FF", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            ws.freeze_panes = "A2"
        print("Saved Excel:", xlsx_path)
    except Exception as e:
        print("Excel error:", e)

    return df

# ======================================================================
# MAIN
# ======================================================================

if __name__ == "__main__":
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 1400)
    pd.set_option("display.float_format", "{:.2f}".format)

    df = apply_thresholds_and_generate_files(TARGET_DIRECTORY, SD_MULTIPLIER)
    print("Rows:", len(df))
